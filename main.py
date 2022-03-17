import openpyxl
from openpyxl.styles import PatternFill, Font
import pandas

book_que= openpyxl.load_workbook("doc_que.xlsx")
sheet=book_que['Sheet1']
round_trips=[]
round_trips_tu=[]
phone_minutes=[]
phone_minutes_tu=[]
row=sheet.max_row
column=sheet.max_column

fill_col1=PatternFill(patternType='solid', fgColor='00ff00')
fill_col2=PatternFill(patternType='solid', fgColor='ff3300')

sheet.cell(15,3,'Suitable')
sheet.cell(15,4).fill=fill_col1
sheet.cell(15,3).font = Font(bold=True)

sheet.cell(16,3,'Not Suitable')
sheet.cell(16,4).fill=fill_col2
sheet.cell(16,3).font = Font(bold=True)

for i in range (2,(row+1)):
    round_trips.append(sheet.cell(i,1).value)
    round_trips_tu.append(sheet.cell(i,2).value)
    phone_minutes.append(sheet.cell(i,3).value)
    phone_minutes_tu.append(sheet.cell(i,4).value)


sheet.cell(1,1,'Round Trips')
sheet.cell(1,2,'Total Utility')
sheet.cell(1,3,'Marginal Utility(Round Trips)')
sheet.cell(1,4,'MU/P $2')
sheet.cell(1,5,'Phone Minutes')
sheet.cell(1,6,'Total Utility')
sheet.cell(1,7,'Marginal Utility(Phone Minutes)')
sheet.cell(1,8,'MU PM $0.05')

sheet.cell(2,1,round_trips[0])
sheet.cell(2,2,round_trips_tu[0])
sheet.cell(2,3,round_trips[0])
sheet.cell(2,4,round_trips[0])
sheet.cell(2,5,phone_minutes[0])
sheet.cell(2,6,phone_minutes_tu[0])
sheet.cell(2,7,phone_minutes[0])
sheet.cell(2,8,phone_minutes[0])

a=[]
b=[]

for i in range(3, (len(round_trips)+2)):
    sheet.cell(i, 1, round_trips[i-2])
    sheet.cell(i, 2, round_trips_tu[i-2])

    MU_round=(round_trips_tu[i-2]-round_trips_tu[i-3])/(round_trips[i-2]-round_trips[i-3])
    PMU_round=MU_round/2
    sheet.cell(i, 3, MU_round)
    sheet.cell(i, 4,PMU_round)
    a.append(PMU_round)

    sheet.cell(i, 5, phone_minutes[i-2])
    sheet.cell(i, 6, phone_minutes_tu[i-2])

    MU_phone_min=(phone_minutes_tu[i-2]-phone_minutes_tu[i-3])/(phone_minutes[i-2]-phone_minutes[i-3])
    PMU_phone_min=MU_phone_min/0.05
    sheet.cell(i,7, MU_phone_min)
    sheet.cell(i,8,PMU_phone_min)
    b.append(PMU_phone_min)


for i in range(0, len(a)):
    for j in range(0, len(b)):
        if (a[i] == b[j]):
            budget=round_trips[i+1]*2+phone_minutes[j+1]*0.05
            if(budget==11):
                sheet.cell(i+3,4).fill=fill_col1
                sheet.cell(j+3,8).fill=fill_col1
                sheet.cell(i+3,4).font=Font(bold=True)
                sheet.cell(j+3,8).font=Font(bold=True)
                print(round_trips[i+1],'* $2 +',phone_minutes[j+1],'* $0.05 = ', budget)
                print("Suitable\n")
            else:
                sheet.cell(i+3,4).fill=fill_col2
                sheet.cell(j+3,8).fill=fill_col2
                sheet.cell(i+3,4).font=Font(bold=True)
                sheet.cell(j+3,8).font=Font(bold=True)
                print(round_trips[i+1],'* $2 +',phone_minutes[j+1],'* $0.05 = ', budget)
                print("Not Suitable\n")
            continue
        else:
            continue

book_que.save('doc_ans.xlsx')

xl_file=pandas.ExcelFile('book_ans.xlsx')
sheet_out=pandas.read_excel(xl_file)
print(sheet_out.head(11))

#Final One